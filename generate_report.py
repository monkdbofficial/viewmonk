from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
import openpyxl

wb = Workbook()

# ─── COLOR PALETTE ────────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"
ACCENT_BLUE = "1565C0"
MID_BLUE    = "1976D2"
LIGHT_BLUE  = "BBDEFB"
HEADER_BG   = "1A237E"
ROW_ALT     = "E8EAF6"
ROW_WHITE   = "FFFFFF"
GOLD        = "F9A825"
GREEN       = "2E7D32"
GREEN_LIGHT = "C8E6C9"
RED         = "C62828"
RED_LIGHT   = "FFCDD2"
ORANGE      = "E65100"
ORANGE_LIGHT= "FFE0B2"
PURPLE      = "6A1B9A"
PURPLE_LIGHT= "E1BEE7"
TEAL        = "00695C"
TEAL_LIGHT  = "B2DFDB"
GRAY_BG     = "F5F5F5"
GRAY_DARK   = "424242"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin"):
    s = Side(style=style, color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(style="medium", color="1A237E")
    return Border(left=t, right=t, top=t, bottom=t)

def set_cell(ws, row, col, value, bold=False, bg=None, fg="000000",
             size=11, h="left", v="center", wrap=False, italic=False, b=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=fg, size=size, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = align(h=h, v=v, wrap=wrap)
    c.border = b if b else border()
    return c

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER PAGE
# ══════════════════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "Cover Page"
ws_cover.sheet_view.showGridLines = False
ws_cover.column_dimensions["A"].width = 4
for col in ["B","C","D","E","F","G","H"]:
    ws_cover.column_dimensions[col].width = 20
for r in range(1, 40):
    ws_cover.row_dimensions[r].height = 22

# Full-width dark banner
for row in range(2, 9):
    for col in range(2, 9):
        c = ws_cover.cell(row=row, column=col)
        c.fill = fill(DARK_NAVY)

ws_cover.merge_cells("B2:H8")
title_cell = ws_cover.cell(row=2, column=2)
title_cell.value = "MONTHLY ACTIVITY REPORT"
title_cell.font = Font(bold=True, color="FFFFFF", size=32, name="Calibri")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
title_cell.fill = fill(DARK_NAVY)

# Gold accent bar
for col in range(2, 9):
    c = ws_cover.cell(row=9, column=col)
    c.fill = fill(GOLD)

# Sub-title band
for row in range(10, 13):
    for col in range(2, 9):
        ws_cover.cell(row=row, column=col).fill = fill(ACCENT_BLUE)

ws_cover.merge_cells("B10:H12")
sub = ws_cover.cell(row=10, column=2)
sub.value = "MonkDB Workbench  |  February 2026"
sub.font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
sub.alignment = Alignment(horizontal="center", vertical="center")
sub.fill = fill(ACCENT_BLUE)

# Details table
details = [
    ("Developer",     "Suryakant Kumar  (suryakant-monkdb)"),
    ("Project",       "MonkDB Workbench"),
    ("Report Month",  "February 2026  (01 Feb – 28 Feb)"),
    ("Total Commits", "31"),
    ("Submitted To",  "Sainath"),
    ("Date",          ""),
    ("Status",        "Pending Review"),
]

row = 15
for label, value in details:
    ws_cover.row_dimensions[row].height = 28
    ws_cover.merge_cells(f"B{row}:C{row}")
    ws_cover.merge_cells(f"D{row}:H{row}")
    lc = ws_cover.cell(row=row, column=2, value=label)
    lc.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    lc.fill = fill(HEADER_BG)
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    vc = ws_cover.cell(row=row, column=4, value=value)
    vc.font = Font(bold=False, color=GRAY_DARK, size=12, name="Calibri")
    vc.fill = fill(ROW_WHITE)
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t = Side(style="thin", color="BDBDBD")
    for col in [2, 4]:
        ws_cover.cell(row=row, column=col).border = Border(left=t, right=t, top=t, bottom=t)
    row += 2

# Footer
ws_cover.merge_cells("B36:H37")
footer = ws_cover.cell(row=36, column=2)
footer.value = "CONFIDENTIAL  |  For Internal Review Only"
footer.font = Font(bold=True, color="FFFFFF", size=10, italic=True, name="Calibri")
footer.fill = fill(DARK_NAVY)
footer.alignment = Alignment(horizontal="center", vertical="center")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — COMMIT LOG
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Commit Log")
ws.sheet_view.showGridLines = False

col_widths = [5, 12, 13, 14, 26, 58, 16, 14, 16]
col_names  = ["#", "Commit ID", "Date", "Category", "Feature Area", "Commit Message", "Files Changed", "Lines Added", "Lines Removed"]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title banner
ws.merge_cells("A1:I1")
ws.row_dimensions[1].height = 36
t = ws.cell(row=1, column=1, value="COMMIT LOG  —  FEBRUARY 2026")
t.font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
t.fill = fill(DARK_NAVY)
t.alignment = Alignment(horizontal="center", vertical="center")

# Gold divider
ws.row_dimensions[2].height = 5
for col in range(1, 10):
    ws.cell(row=2, column=col).fill = fill(GOLD)

# Header row
ws.row_dimensions[3].height = 30
for col, name in enumerate(col_names, 1):
    c = ws.cell(row=3, column=col, value=name)
    c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.fill = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    s = Side(style="medium", color="FFFFFF")
    c.border = Border(left=s, right=s, top=s, bottom=s)

# Category → color mapping
CAT_COLOR = {
    "Feature":      (ACCENT_BLUE, "FFFFFF"),
    "Bug Fix":      (RED,         "FFFFFF"),
    "Cleanup":      (ORANGE,      "FFFFFF"),
    "Refactor":     (PURPLE,      "FFFFFF"),
    "Docs":         (TEAL,        "FFFFFF"),
    "Merge":        (GRAY_DARK,   "FFFFFF"),
    "Build/Release":(GREEN,       "FFFFFF"),
}

commits = [
    (1,  "55aaacd","07-Feb-2026","Feature",      "Schema Security",        "Implement enterprise-grade schema filtering with permission-based isolation",62,5866,1207),
    (2,  "fa57d76","07-Feb-2026","Bug Fix",       "Schema Viewer",          "SchemaViewer schema.name extraction fix for object structure",1,11,10),
    (3,  "56816fe","07-Feb-2026","Bug Fix",       "User Management",        "Use sys.privileges instead of information_schema.table_privileges",3,46,29),
    (4,  "496adfe","07-Feb-2026","Feature",       "Enterprise Features",    "Add enterprise features: ER Diagram, Data Editor, Index Mgmt, User Management",22,9007,0),
    (5,  "fc867c3","08-Feb-2026","Refactor",      "ER Diagram",             "Simplify ER diagram components and add sample data utilities",13,1808,1890),
    (6,  "232ef6c","08-Feb-2026","Feature",       "ER Diagram",             "Add fullscreen and PNG export to SimpleERDiagram",1,191,5),
    (7,  "aaa3dd5","08-Feb-2026","Feature",       "Vector & Search",        "Add Vector Operations, Full-Text Search, and enterprise search features",54,13957,1954),
    (8,  "b3e1cc0","09-Feb-2026","Build/Release", "Build System",           "Add standalone build automation and fix TypeScript build error",5,1406,154),
    (9,  "5ade63f","09-Feb-2026","Merge",         "Branch Management",      "Merge feature/enterprise-schema-filtering into main",0,0,0),
    (10, "92f63c8","09-Feb-2026","Feature",       "Build System",           "Add runtime environment variable support for standalone builds",3,207,14),
    (11, "a94b4e2","09-Feb-2026","Feature",       "Build System",           "Optimize standalone build — exclude node_modules",2,35,9),
    (12, "20df927","09-Feb-2026","Merge",         "Branch Management",      "Merge enterprise-schema-filtering — runtime env + optimized build",0,0,0),
    (13, "dca3cf6","09-Feb-2026","Bug Fix",       "Maps",                   "Fix: Wait for runtime environment to load before initializing Mapbox",1,19,6),
    (14, "814ae97","09-Feb-2026","Bug Fix",       "Maps",                   "Fix: Resolve Mapbox runtime environment race condition",0,0,0),
    (15, "bfb33a3","09-Feb-2026","Build/Release", "Build System",           "Add both LITE and FULL standalone build versions",1,80,27),
    (16, "25a9266","09-Feb-2026","Build/Release", "Build System",           "Add LITE and FULL standalone build options",0,0,0),
    (17, "c618b7b","09-Feb-2026","Feature",       "Maps",                   "Replace Mapbox with Leaflet for zero-dependency maps",4,214,5),
    (18, "dddcb4e","09-Feb-2026","Feature",       "Maps",                   "Complete Leaflet migration and remove Mapbox dependencies",6,249,240),
    (19, "447b678","09-Feb-2026","Merge",         "Branch Management",      "Merge branch feature/enterprise-schema-filtering",0,0,0),
    (20, "ae10f4d","09-Feb-2026","Build/Release", "Release Automation",     "Auto-remove source code from releases",1,21,0),
    (21, "8f07eab","09-Feb-2026","Docs",          "Documentation",          "Add clear warning about which files to download",1,21,14),
    (22, "06a1c88","09-Feb-2026","Docs",          "Documentation",          "Add standalone build download section to README",1,24,1),
    (23, "23f0eec","09-Feb-2026","Build/Release", "Release",                "Publish releases to public repository",1,10,31),
    (24, "2f09ad2","09-Feb-2026","Feature",       "Infrastructure",         "Replace GitHub releases with Cloudflare R2 storage",1,163,85),
    (25, "01f8312","09-Feb-2026","Cleanup",       "Build System",           "Remove public repository dependency",1,0,99),
    (26, "54a61a3","09-Feb-2026","Feature",       "Download Page",          "Add beautiful download page for standalone builds",2,431,0),
    (27, "be14de6","09-Feb-2026","Feature",       "Branding",               "Rebrand to ViewMonk with enterprise-grade download page",2,578,197),
    (28, "644fba8","12-Feb-2026","Feature",       "AQI Module",             "Added the new AQI (Air Quality Index) module",163,2837,8178),
    (29, "794d46b","14-Feb-2026","Feature",       "AQI Module",             "Add setup script for MonkDB AQI Platform — SQL via HTTP API",59,18583,658),
    (30, "421a7f3","23-Feb-2026","Cleanup",       "UI / Layout",            "Remove Index Management page and update layout metadata",31,4637,9405),
    (31, "ca7426b","26-Feb-2026","Feature",       "UI Enhancement",         "Enhance UI and functionality in Unified Browser and Vector Operations",78,6255,25172),
    (32, "8ce8f9d","26-Feb-2026","Cleanup",       "Documentation",          "Remove outdated AQI Solution documentation and presentation files",59,8714,2501),
]

for i, row_data in enumerate(commits):
    r = i + 4
    ws.row_dimensions[r].height = 22
    bg = ROW_ALT if i % 2 == 0 else ROW_WHITE

    sno, cid, date, cat, area, msg, files, added, removed = row_data

    # Serial
    set_cell(ws, r, 1, sno,  bold=True, bg=bg, fg=GRAY_DARK, h="center")
    # Commit ID — monospace style
    c = ws.cell(row=r, column=2, value=cid)
    c.font = Font(bold=True, color=ACCENT_BLUE, size=10, name="Courier New")
    c.fill = fill(bg)
    c.alignment = align(h="center")
    c.border = border()
    # Date
    set_cell(ws, r, 3, date,   bg=bg, h="center", size=10)
    # Category — colored pill
    cat_bg, cat_fg = CAT_COLOR.get(cat, (GRAY_DARK, "FFFFFF"))
    set_cell(ws, r, 4, cat,   bold=True, bg=cat_bg, fg=cat_fg, h="center", size=10)
    # Feature area
    set_cell(ws, r, 5, area,  bg=bg, fg=GRAY_DARK, size=10, wrap=True)
    # Message
    set_cell(ws, r, 6, msg,   bg=bg, fg="212121", size=10, wrap=True)
    # Files
    set_cell(ws, r, 7, files, bg=bg, fg=PURPLE,     h="center", bold=(files>50))
    # Added — green
    set_cell(ws, r, 8, f"+{added:,}", bg=GREEN_LIGHT if added > 0 else bg, fg=GREEN, h="center", bold=True, size=10)
    # Removed — red
    set_cell(ws, r, 9, f"-{removed:,}", bg=RED_LIGHT if removed > 0 else bg, fg=RED, h="center", bold=True, size=10)

# Totals row
r = len(commits) + 4
ws.row_dimensions[r].height = 26
for col in range(1, 10):
    ws.cell(row=r, column=col).fill = fill(DARK_NAVY)
    ws.cell(row=r, column=col).border = border()
ws.merge_cells(f"A{r}:F{r}")
tc = ws.cell(row=r, column=1, value="TOTALS")
tc.font = Font(bold=True, color=GOLD, size=12, name="Calibri")
tc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
tc.fill = fill(DARK_NAVY)

total_files = sum(c[6] for c in commits)
total_added = sum(c[7] for c in commits)
total_removed = sum(c[8] for c in commits)

for col, val, clr in [(7, total_files, GOLD), (8, f"+{total_added:,}", GOLD), (9, f"-{total_removed:,}", GOLD)]:
    c = ws.cell(row=r, column=col, value=val)
    c.font = Font(bold=True, color=clr, size=12, name="Calibri")
    c.fill = fill(DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — WORK SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Work Summary")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 60
ws2.column_dimensions["E"].width = 18

# Banner
ws2.merge_cells("A1:E1")
ws2.row_dimensions[1].height = 36
t = ws2.cell(row=1, column=1, value="WORK SUMMARY BY FEATURE AREA")
t.font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
t.fill = fill(DARK_NAVY)
t.alignment = Alignment(horizontal="center", vertical="center")

ws2.row_dimensions[2].height = 5
for col in range(1, 6):
    ws2.cell(row=2, column=col).fill = fill(GOLD)

# Header
ws2.row_dimensions[3].height = 30
for col, name in enumerate(["#", "Feature Area", "Commits", "Description / Deliverable", "Status"], 1):
    c = ws2.cell(row=3, column=col, value=name)
    c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.fill = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()

work_summary = [
    (1,  "Schema Security & Permissions", 2,  "Enterprise-grade schema filtering with permission-based isolation using sys.privileges instead of information_schema. Fixes privilege queries for correct schema-level access control.",                                           "Completed"),
    (2,  "Enterprise UI Features",         3,  "ER Diagram (with fullscreen + PNG export), Data Editor, Index Management, and User Management pages added as full enterprise-grade features to the workbench sidebar.",                                                       "Completed"),
    (3,  "Vector Operations & FTS",        1,  "New Vector Operations page with KNN search, cosine similarity, collection browser, FLOAT_VECTOR column detection, inline edit/delete, saved searches, analytics panel, and Full-Text Search integration.",                   "Completed"),
    (4,  "Maps — Leaflet Migration",       3,  "Replaced Mapbox (token-dependent) with Leaflet (zero-dependency). Full migration: OpenStreetMap/satellite/topo tiles. GeoPoint CircleMarkers and GeoShape Polygon/LineString rendering via Leaflet.",                         "Completed"),
    (5,  "Standalone Build System",        6,  "Automated LITE and FULL standalone build variants. Runtime environment variable support for MonkDB host config without rebuild. Cloudflare R2 storage for distribution. node_modules excluded from release bundles.",          "Completed"),
    (6,  "Release & Branding",             4,  "Rebranded product to ViewMonk. Added enterprise-grade download page with LITE/FULL build options. GitHub + Cloudflare R2 release automation. Auto-removal of source code from published releases.",                           "Completed"),
    (7,  "AQI Platform",                   2,  "New Air Quality Index (AQI) monitoring module added. Includes setup script for MonkDB AQI Platform executing SQL files via HTTP API. Enables real-time pollution data ingestion and visualization.",                           "Completed"),
    (8,  "UI Enhancements",                2,  "Enhanced Unified Browser and Vector Operations pages with improved UI patterns. Removed Index Management page and updated layout metadata for cleaner navigation.",                                                             "Completed"),
    (9,  "Bug Fixes",                      3,  "Fixed SchemaViewer schema.name object extraction. Corrected sys.privileges query for user management. Resolved Mapbox runtime environment race condition (initialization before env loaded).",                                  "Completed"),
    (10, "Documentation & Cleanup",        2,  "Removed all outdated AQI Solution documentation, presentation slides, quick reference card, README, and setup guides. Added standalone build download instructions and download warnings to README.",                           "Completed"),
]

STATUS_COLOR = {"Completed": (GREEN, GREEN_LIGHT), "In Progress": (ORANGE, ORANGE_LIGHT), "Pending": (RED, RED_LIGHT)}

for i, row_data in enumerate(work_summary):
    r = i + 4
    ws2.row_dimensions[r].height = 55
    bg = ROW_ALT if i % 2 == 0 else ROW_WHITE
    sno, area, commits_count, desc, status = row_data
    set_cell(ws2, r, 1, sno,           bold=True,  bg=bg,        fg=GRAY_DARK, h="center", v="top")
    set_cell(ws2, r, 2, area,          bold=True,  bg=bg,        fg=ACCENT_BLUE, size=11, v="top", wrap=True)
    set_cell(ws2, r, 3, commits_count, bold=True,  bg=LIGHT_BLUE,fg=HEADER_BG,  h="center", v="top", size=13)
    set_cell(ws2, r, 4, desc,          bold=False, bg=bg,        fg="212121",  size=10, wrap=True, v="top")
    s_fg, s_bg = STATUS_COLOR.get(status, (GRAY_DARK, GRAY_BG))
    set_cell(ws2, r, 5, status,        bold=True,  bg=s_bg,      fg=s_fg,      h="center", v="top")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ANALYTICS (charts + stats)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Analytics & Charts")
ws3.sheet_view.showGridLines = False
for col in "ABCDEFGHIJKLMNOP":
    ws3.column_dimensions[col].width = 16

# Banner
ws3.merge_cells("A1:P1")
ws3.row_dimensions[1].height = 36
t = ws3.cell(row=1, column=1, value="ANALYTICS & METRICS  —  FEBRUARY 2026")
t.font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
t.fill = fill(DARK_NAVY)
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[2].height = 5
for col in range(1, 17):
    ws3.cell(row=2, column=col).fill = fill(GOLD)

# ── KPI CARDS (row 4-7) ───────────────────────────────────────────────────────
kpis = [
    ("Total Commits",      "31",        ACCENT_BLUE),
    ("Lines Added",        "74,760",    GREEN),
    ("Lines Removed",      "51,876",    RED),
    ("Net Lines",          "+22,884",   TEAL),
    ("Files Changed",      "577",       PURPLE),
    ("Feature Commits",    "18",        GOLD),
    ("Bug Fixes",          "3",         ORANGE),
    ("Active Days",        "6",         DARK_NAVY),
]

col_start = 1
for kpi_label, kpi_val, kpi_color in kpis:
    ws3.merge_cells(start_row=4, start_column=col_start, end_row=5, end_column=col_start+1)
    ws3.merge_cells(start_row=6, start_column=col_start, end_row=7, end_column=col_start+1)
    lc = ws3.cell(row=4, column=col_start, value=kpi_label)
    lc.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    lc.fill = fill(kpi_color)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws3.cell(row=6, column=col_start, value=kpi_val)
    vc.font = Font(bold=True, color=kpi_color, size=20, name="Calibri")
    vc.fill = fill(ROW_WHITE)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    t1 = Side(style="medium", color=kpi_color)
    for r in [4, 5, 6, 7]:
        for c in [col_start, col_start+1]:
            ws3.cell(row=r, column=c).border = Border(left=t1, right=t1, top=t1, bottom=t1)
    col_start += 2

ws3.row_dimensions[4].height = 20
ws3.row_dimensions[5].height = 20
ws3.row_dimensions[6].height = 35
ws3.row_dimensions[7].height = 20

# ── Category Breakdown Table (data for pie chart) ────────────────────────────
ws3.row_dimensions[9].height = 5
for col in range(1, 9):
    ws3.cell(row=9, column=col).fill = fill(LIGHT_BLUE)

cat_hdr_row = 10
for col, name in enumerate(["Category", "Count", "Share %"], 1):
    c = ws3.cell(row=cat_hdr_row, column=col, value=name)
    c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.fill = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()

categories = [
    ("Feature / New Dev", 18, "56%"),
    ("Bug Fix",            3,  "9%"),
    ("Cleanup",            3, "9%"),
    ("Refactor",           1,  "3%"),
    ("Documentation",      2,  "6%"),
    ("Build / Release",    4, "13%"),
    ("Merge Commits",      3,  "9%"),
]

for i, (cat, cnt, pct) in enumerate(categories):
    r = cat_hdr_row + 1 + i
    bg = ROW_ALT if i % 2 == 0 else ROW_WHITE
    ws3.row_dimensions[r].height = 22
    cat_bg, cat_fg = CAT_COLOR.get(cat.split(" /")[0].strip(), (GRAY_DARK, "FFFFFF"))
    set_cell(ws3, r, 1, cat, bold=True,  bg=cat_bg, fg=cat_fg, h="center")
    set_cell(ws3, r, 2, cnt, bold=True,  bg=bg, fg=GRAY_DARK, h="center")
    set_cell(ws3, r, 3, pct, bold=False, bg=bg, fg=GRAY_DARK, h="center")

# PIE CHART — category breakdown
pie = PieChart()
pie.title = "Commits by Category"
pie.style = 10
pie.width = 14
pie.height = 12
labels_ref = Reference(ws3, min_col=1, min_row=cat_hdr_row+1, max_row=cat_hdr_row+len(categories))
data_ref   = Reference(ws3, min_col=2, min_row=cat_hdr_row,   max_row=cat_hdr_row+len(categories))
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(labels_ref)
pie.dataLabels = openpyxl.chart.label.DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = False
ws3.add_chart(pie, "E10")

# ── Week-wise Activity Table ─────────────────────────────────────────────────
week_start_row = 20
ws3.row_dimensions[week_start_row].height = 5
for col in range(1, 9):
    ws3.cell(row=week_start_row, column=col).fill = fill(LIGHT_BLUE)

for col, name in enumerate(["Week", "Dates", "Commits", "Key Deliverables"], 1):
    c = ws3.cell(row=week_start_row+1, column=col, value=name)
    c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.fill = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()

weeks = [
    ("Week 1", "Feb 07 – 09", 25, "Enterprise features, ER Diagram, Vector Search, Leaflet migration, Standalone build, ViewMonk rebrand"),
    ("Week 2", "Feb 12 – 14",  2, "AQI module development and HTTP API setup script for SQL execution"),
    ("Week 3", "Feb 23 – 26",  4, "Index Management removal, UI enhancements, documentation cleanup"),
]
week_colors = [ACCENT_BLUE, GREEN, ORANGE]
for i, (wk, dates, cnt, key) in enumerate(weeks):
    r = week_start_row + 2 + i
    ws3.row_dimensions[r].height = 40
    set_cell(ws3, r, 1, wk,    bold=True,  bg=week_colors[i], fg="FFFFFF", h="center")
    set_cell(ws3, r, 2, dates, bold=False, bg=ROW_ALT, fg=GRAY_DARK, h="center")
    set_cell(ws3, r, 3, cnt,   bold=True,  bg=LIGHT_BLUE, fg=HEADER_BG, h="center", size=14)
    set_cell(ws3, r, 4, key,   bold=False, bg=ROW_WHITE, fg="212121", wrap=True, size=10)

# BAR CHART — week-wise commits
ws3.cell(row=25, column=1, value="Week")
ws3.cell(row=25, column=2, value="Commits")
for i, (wk, _, cnt, _) in enumerate(weeks):
    ws3.cell(row=26+i, column=1, value=wk)
    ws3.cell(row=26+i, column=2, value=cnt)

bar = BarChart()
bar.type = "col"
bar.style = 10
bar.title = "Commits per Week"
bar.y_axis.title = "Commits"
bar.x_axis.title = "Week"
bar.width = 14
bar.height = 10
bar_data   = Reference(ws3, min_col=2, min_row=25, max_row=28)
bar_labels = Reference(ws3, min_col=1, min_row=26, max_row=28)
bar.add_data(bar_data, titles_from_data=True)
bar.set_categories(bar_labels)
bar.shape = 4
ws3.add_chart(bar, "E22")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SIGN-OFF
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Sign-Off")
ws4.sheet_view.showGridLines = False
for col in "ABCDEFGH":
    ws4.column_dimensions[col].width = 20
for r in range(1, 40):
    ws4.row_dimensions[r].height = 22

# Banner
ws4.merge_cells("A1:H1")
ws4.row_dimensions[1].height = 36
t = ws4.cell(row=1, column=1, value="REPORT SIGN-OFF & REVIEW")
t.font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
t.fill = fill(DARK_NAVY)
t.alignment = Alignment(horizontal="center", vertical="center")

ws4.merge_cells("A2:H2")
ws4.cell(row=2, column=1).fill = fill(GOLD)

def sign_section(ws, start_row, title, fields):
    ws.merge_cells(f"A{start_row}:H{start_row}")
    tc = ws.cell(row=start_row, column=1, value=title)
    tc.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    tc.fill = fill(ACCENT_BLUE)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r = start_row + 1
    for label, default in fields:
        ws.merge_cells(f"A{r}:C{r}")
        ws.merge_cells(f"D{r}:H{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, color=GRAY_DARK, size=11, name="Calibri")
        lc.fill = fill(GRAY_BG)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc = ws.cell(row=r, column=4, value=default)
        vc.font = Font(bold=False, color=GRAY_DARK, size=11, name="Calibri")
        vc.fill = fill(ROW_WHITE)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        s = Side(style="thin", color="BDBDBD")
        b = Border(left=s, right=s, top=s, bottom=s)
        for col in [1, 4]:
            ws.cell(row=r, column=col).border = b
        ws.row_dimensions[r].height = 28
        r += 1
    return r + 1

r = sign_section(ws4, 4, "SUBMITTED BY", [
    ("Name",            "Suryakant Kumar"),
    ("Role",            "Developer"),
    ("Date Submitted",  ""),
    ("Signature",       ""),
])

r = sign_section(ws4, r, "REVIEWED BY — SAINATH", [
    ("Reviewer Name",   "Sainath"),
    ("Review Date",     ""),
    ("Observations",    ""),
    ("Comments",        ""),
    ("Decision",        "Approved  /  Revisions Needed  /  Pending"),
    ("Signature",       ""),
])

# Footer
ws4.merge_cells(f"A{r+2}:H{r+3}")
footer = ws4.cell(row=r+2, column=1)
footer.value = "CONFIDENTIAL  |  MonkDB Workbench  |  Internal Report  |  February 2026"
footer.font = Font(bold=True, color="FFFFFF", size=10, italic=True, name="Calibri")
footer.fill = fill(DARK_NAVY)
footer.alignment = Alignment(horizontal="center", vertical="center")


# ── Save ──────────────────────────────────────────────────────────────────────
output = "/Users/surykantkumar/Development/monkdb/workbanch/Monthly_Report_Feb2026_MonkDB.xlsx"
wb.save(output)
print(f"Saved: {output}")
